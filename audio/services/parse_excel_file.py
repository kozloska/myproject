import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import logging
from django.db import transaction
from django.utils import timezone
from sympy import false

from ..models import DefenseSchedule, Specialization, Group, Project, Student, Protocol

logger = logging.getLogger(__name__)
from django.db.models import Q
import pandas as pd
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import logging
from django.db import transaction
from django.utils import timezone
from ..models import DefenseSchedule, Specialization, Group, Project, Student, Protocol

logger = logging.getLogger(__name__)

import pandas as pd
from datetime import datetime
import logging
from django.db import transaction
from ..models import Specialization, Group, Project, Student, Protocol


def parse_excel_file(file_path, specialization_id, institute_id=None):
    """
    Парсит Excel-файл и сохраняет данные в базу данных, обеспечивая единичные записи для проектов и групп.
    Студенты создаются, даже если у них нет отчества (Patronymic=None).
    Поле Supervisor в Project обновляется из столбца 'Преподаватель' для каждой строки.
    Столбец 'Преподаватель' обязателен.
    Для каждого студента создаётся Protocol с ID_DefenseSchedule=None, Year=текущий год, Status=False.
    Поле Protocol.Number формируется как '<порядковый номер за запуск для специализации><буквы из Specialization.Number>',
    сбрасываясь до 1 в начале каждого нового года и запуска.

    Args:
        file_path (str): Путь к Excel-файлу.
        specialization_id (int): ID направления (Specialization).
        institute_id (int, optional): Не используется в текущей версии, но сохранен для совместимости.

    Returns:
        dict: Результат обработки (успех/ошибка, количество добавленных записей).
    """
    try:
        # Читаем Excel-файл
        df = pd.read_excel(file_path, sheet_name=0)

        # Проверяем наличие необходимых колонок
        required_columns = ['ФИО', 'Группа', 'Тема проекта', 'Преподаватель']
        if not all(col in df.columns for col in required_columns):
            logger.error(f"Excel file is missing required columns: {', '.join(required_columns)}")
            return {"status": "error", "message": f"Missing required columns: {', '.join(required_columns)}"}

        # Проверяем существование Specialization
        try:
            specialization = Specialization.objects.get(ID=specialization_id)
            logger.debug(f"Using specialization ID: {specialization_id}, Number: {specialization.Number}")
        except Specialization.DoesNotExist:
            logger.error(f"Specialization with ID {specialization_id} not found")
            return {"status": "error", "message": f"Invalid specialization ID {specialization_id}"}

        # Получаем буквы из поля Number специализации
        letters = specialization.Number or ""  # Используем пустую строку, если букв нет
        if not letters:
            logger.warning(f"No letters found in Specialization.Number for specialization {specialization_id}, using empty letters")

        groups_added = 0
        students_added = 0
        projects_added = 0
        projects_updated = 0
        protocols_added = 0

        # Атомарная транзакция для целостности данных
        with transaction.atomic():
            # Счётчик протоколов для текущего запуска
            protocol_counter = 0

            for index, row in df.iterrows():
                # Пропускаем строки с пустыми ФИО или Группа
                if pd.isna(row['ФИО']) or pd.isna(row['Группа']):
                    logger.debug(f"Skipping row {index + 2}: Empty ФИО or Группа")
                    continue

                # Проверяем тему проекта
                project_title = str(row.get('Тема проекта', '')).strip()
                if not project_title or project_title.lower() == 'nan':
                    logger.debug(f"Skipping row {index + 2}: Empty project title")
                    continue

                # Пропускаем отчисленных студентов
                if 'ОТЧИС' in project_title.upper():
                    logger.debug(f"Skipping row {index + 2}: Student marked as 'ОТЧИСЛЕН'")
                    continue

                # Извлекаем название группы
                group_name = str(row['Группа']).strip()

                # Явная проверка существования группы
                if Group.objects.filter(Name=group_name).exists():
                    logger.info(f"Group '{group_name}' already exists, using existing record")
                    group = Group.objects.get(Name=group_name)
                else:
                    group = Group.objects.create(Name=group_name)
                    groups_added += 1
                    logger.debug(f"Created group: {group_name}")

                # Парсим ФИО
                fio = str(row['ФИО']).strip().split()
                if len(fio) < 2:
                    logger.warning(f"Invalid FIO format at row {index + 2}: {row['ФИО']}")
                    continue
                surname = fio[0]
                name = fio[1]
                patronymic = " ".join(fio[2:]) if len(fio) > 2 else None

                # Извлекаем руководителя
                supervisor = str(row['Преподаватель']).strip()
                supervisor = None if not supervisor or supervisor.lower() == 'nan' else supervisor

                # Находим или создаем проект, обновляем Supervisor
                try:
                    project = Project.objects.get(Title=project_title)
                    # Обновляем Supervisor, если он отличается
                    if project.Supervisor != supervisor:
                        project.Supervisor = supervisor
                        project.save()
                        projects_updated += 1
                        logger.debug(f"Updated project Supervisor: {project_title}, Supervisor: {supervisor or 'None'}")
                except Project.DoesNotExist:
                    project = Project.objects.create(
                        Title=project_title,
                        Supervisor=supervisor,
                        Status='Защита не начата',
                        Text=''
                    )
                    projects_added += 1
                    logger.debug(f"Created project: {project_title}, Supervisor: {supervisor or 'None'}")

                # Создаем студента (уникальность по ФИО, группе и специализации)
                try:
                    student, created = Student.objects.get_or_create(
                        Surname=surname,
                        Name=name,
                        Patronymic=patronymic,
                        ID_Group=group,
                        ID_Specialization=specialization,
                        defaults={'ID_Project': project}
                    )
                    if created:
                        students_added += 1
                        logger.debug(f"Created student: {surname} {name} {patronymic or ''}")
                except Exception as e:
                    logger.error(f"Error creating student at row {index + 2}: {str(e)}")
                    continue

                # Обновляем связь студента с проектом, если она изменилась
                if student.ID_Project != project:
                    student.ID_Project = project
                    student.save()

                # Создаем Protocol для студента, если его ещё нет с ID_DefenseSchedule=None
                if not Protocol.objects.filter(ID_Student=student, ID_DefenseSchedule=None).exists():
                    # Увеличиваем счётчик для текущего запуска
                    protocol_counter += 1

                    # Формируем номер протокола: <порядковый номер><буквы>
                    protocol_number_str = f"{protocol_counter}{letters}"

                    # Создаём протокол с номером
                    protocol = Protocol.objects.create(
                        ID_Student=student,
                        Year=datetime.now().year,
                        Status=False,
                        Grade=None,
                        ID_Question=None,
                        ID_Question2=None,
                        DefenseStartTime=None,
                        DefenseEndTime=None,
                        Number=protocol_number_str
                    )
                    protocols_added += 1
                    logger.debug(f"Created Protocol for student: {student}, Number: {protocol_number_str}")
                else:
                    logger.debug(f"Protocol already exists for student: {student}")

        logger.info(f"Parsed Excel: {groups_added} groups, {students_added} students, {projects_added} projects added, {projects_updated} projects updated, {protocols_added} protocols added")
        return {
            "status": "success",
            "Групп добавлено": groups_added,
            "Студентов создано": students_added,
            "Проектов добавлено": projects_added,
            "Проектов обновлено": projects_updated,
            "Протоколов добавлено": protocols_added
        }

    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        return {"status": "error", "message": str(e)}




def parse_defense_schedule(file_path):
    """
    Парсит Excel-файл с расписанием защит, создавая новые записи в DefenseSchedule и обновляя существующие Protocol.
    Использует существующую Specialization, парсит даты, время и аудиторию из объединённой ячейки.
    Создаёт новую запись DefenseSchedule на временной слот, с Count равным количеству строк, объединённых в столбце B (аудитория).
    Обновляет существующие протоколы, привязывая их к новому DefenseSchedule, даже если они уже связаны с другим.
    Добавляет аудиторию в поле Class. Извлекает все проекты из столбца D в пределах объединённых строк.

    Args:
        file_path (str): Путь к Excel-файлу.

    Returns:
        dict: Результат обработки (успех/ошибка, количество добавленных и связанных записей).
    """
    try:
        # Загружаем Excel-файл с помощью openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        defenses_added = 0
        protocols_linked = 0

        with transaction.atomic():
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")

                # Первая строка — название специальности
                specialization_name = str(sheet['A1'].value).strip() if sheet['A1'].value else None
                if not specialization_name or specialization_name.lower() == 'nan':
                    logger.error(f"No specialization name found in sheet {sheet_name}")
                    continue

                # Находим существующую Specialization
                try:
                    specialization = Specialization.objects.get(Name=specialization_name)
                except Specialization.DoesNotExist:
                    logger.error(f"Specialization '{specialization_name}' not found in database")
                    continue
                logger.debug(f"Using specialization: {specialization_name}")

                # Парсим строки для поиска дат и таблиц
                current_date = None
                current_time_range = None
                current_auditorium = None
                table_rows = []
                in_table = False
                merged_row_count = 0
                table_start_row = None
                time_row_found = False

                # Проходим по строкам листа
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Отладка: выводим содержимое строки
                    logger.debug(f"Row {row_idx}: {list(row)}")

                    # Ищем строку с датой
                    row_str = str(row[0]).strip() if row[0] else ''
                    date_match = re.match(r'(\d{2}\.\d{2}\.\d{4})(?:\s*-\s*\d{2}\.\d{2}\.\d{4})?\s*-\s*[^\s]+', row_str)
                    if date_match:
                        # Если уже есть собранные строки таблицы, создаём DefenseSchedule
                        if table_rows and current_date and current_time_range:
                            slot_count = merged_row_count  # Количество объединённых строк в столбце B
                            time_match = re.match(r'(\d{2}:\d{2})-(\d{2}:\d{2})', current_time_range)
                            if time_match:
                                start_time_str, end_time_str = time_match.groups()
                                try:
                                    start_time = datetime.strptime(start_time_str, '%H:%M').time()
                                except ValueError as e:
                                    logger.error(f"Invalid time format for DefenseSchedule: {current_time_range}, error: {str(e)}")
                                    table_rows = []
                                    merged_row_count = 0
                                    in_table = False
                                    table_start_row = None
                                    time_row_found = False
                                    continue

                                # Создаём новую запись DefenseSchedule
                                defense_datetime = timezone.make_aware(datetime.combine(current_date, start_time))
                                defense_schedule = DefenseSchedule.objects.create(
                                    DateTime=defense_datetime,
                                    ID_Commission=None,
                                    Count=slot_count,
                                    Class=current_auditorium
                                )
                                defenses_added += 1
                                logger.debug(f"Created DefenseSchedule: {defense_datetime}, Count: {slot_count}, Class: {current_auditorium}")

                                # Связываем протоколы для всех проектов в слоте, обновляя существующие
                                for group_name, project_title in table_rows:
                                    if not group_name or not project_title:
                                        continue
                                    try:
                                        group = Group.objects.get(Name=str(group_name).strip())
                                        project = Project.objects.get(Title=str(project_title).strip())
                                    except (Group.DoesNotExist, Project.DoesNotExist) as e:
                                        logger.warning(f"Group or Project not found for {group_name}, {project_title}: {str(e)}")
                                        continue

                                    students = Student.objects.filter(
                                        ID_Project=project,
                                        ID_Group=group,
                                        ID_Specialization=specialization
                                    )
                                    for student in students:
                                        protocols = Protocol.objects.filter(ID_Student=student)
                                        logger.debug(f"Found {protocols.count()} protocols for student {student}")
                                        for protocol in protocols:
                                            protocol.ID_DefenseSchedule = defense_schedule  # Обновляем связь
                                            protocol.save()
                                            protocols_linked += 1
                                            logger.debug(f"Updated Protocol {protocol.ID} to DefenseSchedule {defense_schedule.ID} for student {student}")
                            else:
                                logger.error(f"Time range not matched for {current_time_range}")

                        # Обновляем текущую дату
                        date_str = date_match.group(1)
                        try:
                            current_date = datetime.strptime(date_str, '%d.%m.%Y')
                            logger.debug(f"Found date: {date_str}")
                            table_rows = []
                            merged_row_count = 0
                            current_time_range = None
                            current_auditorium = None
                            in_table = False
                            table_start_row = None
                            time_row_found = False
                        except ValueError:
                            logger.error(f"Invalid date format at row {row_idx}: {row_str}")
                        continue

                    # Ищем заголовок таблицы
                    if row[0] == 'время' and row[1] == 'Аудитория' and row[2] == 'группа' and row[3] == 'тема проекта':
                        logger.debug(f"Found table header at row {row_idx}")
                        in_table = True
                        table_start_row = row_idx + 1  # Следующая строка после заголовка
                        continue

                    # Ищем строку с временем и извлекаем проект
                    if in_table and row_idx == table_start_row and not time_row_found:
                        time_str = str(row[0]).strip() if row[0] else ''
                        time_match = re.match(r'(\d{2}:\d{2})-(\d{2}:\d{2})', time_str)
                        if time_match:
                            start_time_str, end_time_str = time_match.groups()
                            current_time_range = f"{start_time_str}-{end_time_str}"
                            current_auditorium = str(row[1]).strip() if row[1] else None
                            logger.debug(f"Found time range: {current_time_range}, auditorium: {current_auditorium}")

                            # Извлекаем группу и проект из строки с временем
                            group_name = str(row[2]).strip() if row[2] else ''
                            project_title = str(row[3]).strip() if row[3] else ''
                            if project_title:  # Добавляем только если есть проект
                                table_rows.append((group_name, project_title))
                                logger.debug(f"Added table row from time row: {group_name}, {project_title}")

                            # Определяем количество объединённых строк в столбце B
                            merged_row_count = 0
                            for merged_range in sheet.merged_cells.ranges:
                                if merged_range.min_col == 2 and merged_range.max_col == 2:
                                    if merged_range.min_row == row_idx:
                                        merged_row_count = merged_range.max_row - merged_range.min_row + 1
                                        logger.debug(f"Merged rows in column B at row {row_idx}: {merged_row_count}")
                                        break
                            if merged_row_count == 0:
                                logger.warning(f"No merged cells found in column B at row {row_idx}, defaulting to 1")
                                merged_row_count = 1
                            time_row_found = True
                            table_start_row = row_idx + 1  # Данные начинаются со следующей строки
                            continue

                    # Обрабатываем все строки таблицы в пределах merged_row_count
                    if in_table and row_idx >= table_start_row and row_idx < table_start_row + merged_row_count:
                        group_name = str(row[2]).strip() if row[2] else ''
                        project_title = str(row[3]).strip() if row[3] else ''
                        if project_title:  # Добавляем только если есть проект
                            table_rows.append((group_name, project_title))
                            logger.debug(f"Added table row: {group_name}, {project_title}")
                        else:
                            logger.debug(f"Empty table row at row {row_idx}")
                    elif in_table and row_idx >= table_start_row + merged_row_count:
                        in_table = False
                        table_start_row = None
                        time_row_found = False

                # Обработка последней таблицы после завершения цикла
                if current_date and current_time_range:
                    slot_count = merged_row_count  # Количество объединённых строк в столбце B
                    time_match = re.match(r'(\d{2}:\d{2})-(\d{2}:\d{2})', current_time_range)
                    if time_match:
                        start_time_str, end_time_str = time_match.groups()
                        try:
                            start_time = datetime.strptime(start_time_str, '%H:%M').time()
                        except ValueError as e:
                            logger.error(f"Invalid time format for DefenseSchedule: {current_time_range}, error: {str(e)}")
                            continue

                        # Создаём новую запись DefenseSchedule
                        defense_datetime = timezone.make_aware(datetime.combine(current_date, start_time))
                        defense_schedule = DefenseSchedule.objects.create(
                            DateTime=defense_datetime,
                            ID_Commission=None,
                            Count=slot_count,
                            Class=current_auditorium
                        )
                        defenses_added += 1
                        logger.debug(f"Created DefenseSchedule: {defense_datetime}, Count: {slot_count}, Class: {current_auditorium}")

                        # Связываем протоколы для всех проектов в слоте, обновляя существующие
                        for group_name, project_title in table_rows:
                            if not group_name or not project_title:
                                continue
                            try:
                                group = Group.objects.get(Name=str(group_name).strip())
                                project = Project.objects.get(Title=str(project_title).strip())
                            except (Group.DoesNotExist, Project.DoesNotExist) as e:
                                logger.warning(f"Group or Project not found for {group_name}, {project_title}: {str(e)}")
                                continue

                            students = Student.objects.filter(
                                ID_Project=project,
                                ID_Group=group,
                                ID_Specialization=specialization
                            )
                            for student in students:
                                protocols = Protocol.objects.filter(ID_Student=student, Status=false)
                                logger.debug(f"Found {protocols.count()} protocols for student {student}")
                                for protocol in protocols:
                                    protocol.ID_DefenseSchedule = defense_schedule  # Обновляем связь
                                    protocol.save()
                                    protocols_linked += 1
                                    logger.debug(f"Updated Protocol {protocol.ID} to DefenseSchedule {defense_schedule.ID} for student {student}")
                    else:
                        logger.error(f"Time range not matched for {current_time_range} in final table")

        logger.info(f"Parsed defense schedule: {defenses_added} defenses added, {protocols_linked} protocols linked")
        return {
            "status": "success",
            "defenses_added": defenses_added,
            "protocols_linked": protocols_linked
        }

    except Exception as e:
        logger.error(f"Error parsing defense schedule Excel: {str(e)}", exc_info=True)
        return {"status": "error", "message": str(e)}